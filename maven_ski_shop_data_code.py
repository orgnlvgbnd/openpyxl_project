import time
import openpyxl as xl
import string
workbook = xl.load_workbook(filename =r'C:\Users\sehga\OneDrive\Desktop\maven_ski_shop_project.xlsx')
orders = workbook['Orders_Info']
max_row = orders.max_row
max_column = orders.max_column

#PART1
#q1 Column Printer Function
def column_printer(worksheet, column):
    for tup in enumerate(worksheet[column], start = 1):
        print(f'{column.upper()}{tup[0]}',tup[1].value)
        print('----------')
# Works okay!!!
print('Column Printer Function')
column_printer(orders,'b')
#this function has no impact on the max_row and max_column values


#q2 Order data Dictionary
#Create a dictionary with all of the information contained in the 'Orders_Info' Worksheet.
mapping = dict(enumerate(string.ascii_uppercase, start = 1))
orders_dict = {}
for i in range(2,29):
    row_lst = []
    for j in range(1,9):
        if j == 1:
            cell_key = orders[f'{mapping[j]}{i}'].value
        elif j == 5 or j ==6:
            pass
        elif j == 8:
            cell8 = [int(x) for x in str((orders[f'{mapping[j]}{i}'].value)).split(',')]
            row_lst.append(cell8)
        else:
            row_lst.append(orders[f'{mapping[j]}{i}'].value)
    orders_dict.update({cell_key:row_lst})
print(orders_dict)
print()
#Works okay!!!

#q3. Sales tax calculation
#We need to calculate the sales tax and total amount owed for every order in this sheet.
#If location is Sun Valley, apply a sales tax of 8%
#If location is Mammoth, apply a sales tax of 7.75%.
#If location is Stowe, apply a sales tax of 6%
#Use the tax_calculator function to apply sales tax to each subtotal.

for tup in enumerate((orders['d']),start = 1):

    if orders[f'g{tup[0]}'].value == 'Sun Valley':
        rate = 0.08
    elif orders[f'g{tup[0]}'].value == 'Mammoth':
        rate = 0.0775
    else:
        rate = 0.06    
    if tup[0] <= 28:
        if tup[0] == 1:
            pass
        else:
            orders[f'e{tup[0]}'] = tup[1].value * (rate)
            orders[f'f{tup[0]}'] = tup[1].value * (rate+1)
            key = orders[f'a{tup[0]}'].value
            
            orders_dict[key].insert(3,round(orders[f'e{tup[0]}'].value,2))
            orders_dict[key].insert(4,round(orders[f'f{tup[0]}'].value,2))
#Works okay!!!
print(orders_dict)


#  PART 2


# q1. This functions totals the subtotal, taxes, total columns in the orders worksheet
def total_column(col_name):
    total = 0
    def excel_col_name_number(char):
        number = 0
        for i in char:
            number = number + (ord(i)-ord('a')+1)
        return number
    
    for tup in enumerate(orders[col_name], start = 1):
        if tup[0] == 1:
            pass 
        else:
            total += tup[1].value
        if tup[0] == 28:
            break 
              
    orders.cell(row=tup[0]+1, column=excel_col_name_number(col_name), value= total)
for col in ['d','e','f']:
    total_column(col)
workbook.save(filename =r'C:\Users\sehga\OneDrive\Desktop\maven_ski_shop_project.xlsx')
#Works okay!!!

# Average of sub totals
total = 0
count = 0
for value_lst in orders_dict.values():
    total += value_lst[2]
    count += 1
average_sub_tot = total/count
print(average_sub_tot)
#Works Okay!!!

#Finding unique customers and number of orders per customer
total_customers = list()
for value_lst in orders_dict.values():
    total_customers.append(value_lst[0])
#printing number of unique customers
print(len(set(total_customers)))
#Works okay!!!

#finding number of orders per customer
cust_order_dict = {}
cust_seen=[]
for ele in total_customers:
    if ele not in cust_seen:
        cust_order_dict.update({ele:total_customers.count(ele)})
        cust_seen.append(ele)
print(cust_order_dict)
#Works okay!!!

# finding the total number of items sold
total_items_sold = 0
for value_list in orders_dict.values():
    total_items_sold += len(value_list[6])
print(total_items_sold)
time.sleep(0)
# Finding sales by location
unique_locations_dict = { lst[5]:0 for lst in orders_dict.values()}
print(unique_locations_dict)
for location in unique_locations_dict.keys():
    total_location_sale = 0
    for list_values in orders_dict.values():
        if list_values[5] == location:
            total_location_sale += list_values[2]
    unique_locations_dict[location] = round(total_location_sale,2)
print(unique_locations_dict)
#Works okay!!!


#challenge : Aggregator function
def aggregator_function(agg_value_col_no ,agg_key_col_no, dictionary):
    unique_key_dict={lst[agg_key_col_no - 2]:0 for lst in dictionary.values()}
    print(unique_key_dict)
    
    for key in unique_key_dict.keys():
        agg_value = 0
        for list_values in dictionary.values():
            if list_values[agg_key_col_no - 2] == key:
                agg_value += list_values[agg_value_col_no-2]
        unique_key_dict[key] = round(agg_value,2)
    return unique_key_dict

test_v = aggregator_function(4,3,orders_dict)
print(test_v)

#Works Okay!!!